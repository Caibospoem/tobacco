"""
读取 data/ 目录下最新的 .xlsx 文件，调用 Claude API 解析，
输出 data/parsed_strategy.json

JSON 格式:
[
  {
    "code": "13320189",
    "brand": "利群(休闲金中支)",
    "category": "一档",
    "region": "全区",
    "retailers": 12037,
    "tiers": {"30": 1, "29": 1, "28": 1, ...}  // 只含非零挡位
  }
]
"""

import os
import json
import glob
import anthropic

def find_latest_xlsx():
    files = glob.glob("data/*.xlsx")
    if not files:
        print("未找到 xlsx 文件")
        exit(1)
    return max(files, key=os.path.getmtime)


def read_xlsx(filepath):
    """读取 xlsx 全部数据为文本，供 AI 理解"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 400),
                               max_col=36, values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        rows.append("\t".join(cells))

    return "\n".join(rows)


def parse_with_ai(text):
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    prompt = f"""你是一个数据处理助手。以下是福建烟草全区卷烟货源投放策略表的原始数据（TSV格式）。

表格结构：
- 前4行是标题和表头
- 列A: 商品代码
- 列B: 品牌名称
- 列C: 品类（一档~八档）
- 列D: 区域
- 列E: 投放户数
- 列F-AI: 30档~1档的配额数量（条）
- 列AJ: 备注（投放/不投/新品）

请提取所有有效商品（有8位数字代码的行），输出严格 JSON 数组，每个商品一个对象：

JSON 字段：
- code: 商品代码（字符串）
- brand: 品牌名称
- category: 品类
- region: 区域
- retailers: 投放户数（整数）
- quotas: 对象，key是档位数字，value是配额（只含>0的档位），例如 {{"30":1,"29":1}}

只输出 JSON 数组，不要任何其他文字。

原始数据：
{text[:80000]}
"""

    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8000,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.content[0].text


def main():
    xlsx_path = find_latest_xlsx()
    print(f"解析: {xlsx_path}")

    text = read_xlsx(xlsx_path)
    print(f"读取: {len(text)} 字符")

    result = parse_with_ai(text)
    # 提取 JSON（去掉可能的 markdown 包裹）
    start = result.find("[")
    end = result.rfind("]") + 1
    if start >= 0 and end > start:
        result = result[start:end]

    data = json.loads(result)
    print(f"解析完成: {len(data)} 条商品")

    os.makedirs("data", exist_ok=True)
    with open("data/parsed_strategy.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("已保存: data/parsed_strategy.json")


if __name__ == "__main__":
    main()
